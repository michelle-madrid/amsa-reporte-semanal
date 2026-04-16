"""Funciones para seleccionar archivos, leer Excel y exportar rangos como imágenes."""

import os
import time
import tkinter as tk
from tkinter import filedialog

import openpyxl
import win32com.client as win32
from openpyxl.utils import get_column_letter, column_index_from_string
from PIL import ImageGrab

import state
from config import SSO_MARCADOR_TABLA

# Crea una ventana raíz oculta para usar los diálogos de selección sin mostrar una ventana principal.
def _crear_root_oculto():
    root = tk.Tk()
    root.attributes('-topmost', True)
    root.withdraw()
    return root

# Abre un selector para escoger un archivo.
def seleccionar_archivo(mensaje):
    print(f"\nSelecciona {mensaje}:")
    root = _crear_root_oculto()
    try:
        archivo = filedialog.askopenfilename(title=f"Selecciona {mensaje}")
    finally:
        root.destroy()
    if not archivo:
        state.errores.append(f"[ERROR] No se seleccionó {mensaje}")
    return archivo

# Abre un selector para escoger una carpeta.
def seleccionar_carpeta():
    root = _crear_root_oculto()
    try:
        carpeta = filedialog.askdirectory(title="Selecciona carpeta para guardar el informe")
    finally:
        root.destroy()
    return carpeta

# Obtiene o crea una instancia reutilizable de Excel por COM.
def _obtener_excel_app():
    if state._excel_app is None:
        for intento in range(5):
            try:
                # DispatchEx crea siempre un proceso Excel propio (no se conecta a uno existente),
                # lo que garantiza el modelo de objetos completo sin depender del gencache.
                state._excel_app = win32.DispatchEx("Excel.Application")
                state._excel_app.Visible = True
                state._excel_app.DisplayAlerts = False
                break
            except Exception as e:
                if intento < 4:
                    time.sleep(2)
                else:
                    raise RuntimeError(f"No se pudo iniciar Excel después de 5 intentos: {e}")
    return state._excel_app

# Cierra todos los libros abiertos y finaliza la instancia de Excel.
def cerrar_excels():
    if state._excel_app is not None:
        try:
            for wb in list(state._excel_app.Workbooks):
                wb.Close(False)
            state._excel_app.Quit()
        except Exception:
            pass
        state._excel_app = None
        state._workbooks_abiertos = {}

# Exporta un rango de Excel como imagen.
def exportar_imagen_excel(ruta_excel, hoja, rango, nombre_imagen):
    return exportar_imagen_excel_rangos(ruta_excel, hoja, [rango], nombre_imagen)

# Exporta uno o varios rangos de Excel como una sola imagen.
def exportar_imagen_excel_rangos(ruta_excel, hoja, lista_rangos, nombre_imagen):
    """
    Copia una o más áreas de Excel como imagen. Si hay varios rangos, usa Union
    (útil para repetir fila de encabezado + cuerpo).
    """
    if not lista_rangos:
        state.errores.append(f"[ERROR] Sin rangos para exportar ({nombre_imagen})")
        return os.path.join(r"C:\\Temp", nombre_imagen)

    carpeta_temp = r"C:\\Temp"
    if not os.path.exists(carpeta_temp):
        os.makedirs(carpeta_temp)

    imagen_salida = os.path.join(carpeta_temp, nombre_imagen)
    rango_desc = ",".join(lista_rangos)

    try:
        excel = _obtener_excel_app()
        wb = state._workbooks_abiertos.get(ruta_excel)
        if wb is None:
            wb = excel.Workbooks.Open(ruta_excel, UpdateLinks=0)
            state._workbooks_abiertos[ruta_excel] = wb

        ws = wb.Worksheets(hoja)
        rng = ws.Range(lista_rangos[0])
        for addr in lista_rangos[1:]:
            rng = excel.Union(rng, ws.Range(addr))

        rng.CopyPicture(Appearance=1, Format=2)
        time.sleep(1)
        img = ImageGrab.grabclipboard()
        if img:
            img.save(imagen_salida, "PNG")
        else:
            msg = f"[ERROR] No se pudo obtener imagen del portapapeles ({hoja} {rango_desc})"
            state.errores.append(msg)
            print(f"  ✗ {msg}")

    except Exception as e:
        msg = f"[ERROR] Falló exportación de imagen {hoja} {rango_desc}: {e}"
        state.errores.append(msg)
        print(f"  ✗ {msg}")

    return imagen_salida

# Implementa una parte específica de la lógica del informe.
def _fila_tiene_contenido_util(row_vals):
    """True si la fila no es solo vacíos y ceros (para recortar relleno al final)."""
    for v in row_vals:
        if v is None:
            continue
        if isinstance(v, (int, float)):
            if v != 0:
                return True
            continue
        if isinstance(v, str):
            s = v.strip()
            if s == "":
                continue
            try:
                if float(s.replace(",", ".")) != 0:
                    return True
            except ValueError:
                return True
            continue
        return True
    return False

# Implementa una parte específica de la lógica del informe.
def _ultima_fila_con_datos_en_rango_com(ws_com, min_col, min_row, max_col, max_row):
    """Usa la hoja ya abierta en Excel (COM) para no bloquear el archivo con otro lector."""
    for r in range(max_row, min_row - 1, -1):
        vals = [ws_com.Cells(r, c).Value for c in range(min_col, max_col + 1)]
        if _fila_tiene_contenido_util(vals):
            return r
    return min_row



def _columna_izquierda_tabla_sso(ws_com, fila_encabezado, max_col_limit=40):
    """Primera columna con texto en la fila de encabezado (por si hay celdas combinadas)."""
    for c in range(1, max_col_limit + 1):
        v = ws_com.Cells(fila_encabezado, c).Value
        if v is not None and str(v).strip() != "":
            return c
    return 1

# Implementa una parte específica de la lógica del informe.
def _ultima_columna_cabecera_sso(ws_com, fila_encabezado, max_col_limit=40):
    """Última columna con texto en la fila del encabezado de la tabla."""
    last = 1
    for c in range(1, max_col_limit + 1):
        v = ws_com.Cells(fila_encabezado, c).Value
        if v is not None and str(v).strip() != "":
            last = c
    return max(last, 1)

# Implementa una parte específica de la lógica del informe.
def _filas_encabezado_tablas_sso(ws_com, marcador=SSO_MARCADOR_TABLA):
    """
    Devuelve lista de (fila, col_marcador) donde está el encabezado de cada tabla SSO.
    Busca el marcador en las primeras 20 columnas.
    """
    used = ws_com.UsedRange
    max_r = used.Row + used.Rows.Count - 1
    filas = []
    for r in range(1, max_r + 1):
        for c in range(1, 21):
            v = ws_com.Cells(r, c).Value
            if v is None:
                continue
            if marcador in str(v).strip().lower():
                filas.append((r, c))
                break
    if not filas:
        msg = f"[REVISAR] SSO: no se encontró el marcador '{marcador}' en la hoja. Verifica que la columna 'Id del incidente' exista."
        state.errores.append(msg)
        print(f"  ! {msg}")
    return filas

# Verifica si una tabla SSO tiene al menos un Id del incidente distinto de 0.
def _tabla_sso_tiene_datos(ws_com, h_row, last_row, min_c, last_c):
    """Busca la columna 'Id del incidente' y omite la tabla si todos los IDs son 0 o None."""
    col_id = None
    for c in range(min_c, last_c + 1):
        v = ws_com.Cells(h_row, c).Value
        if v is not None and SSO_MARCADOR_TABLA in str(v).strip().lower():
            col_id = c
            break

    if col_id is None:
        return True  # No encontramos la columna, incluir por precaución

    for r in range(h_row + 1, last_row + 1):
        v = ws_com.Cells(r, col_id).Value
        if v is not None and v != 0:
            return True

    return False

# Implementa una parte específica de la lógica del informe.
def _rangos_tablas_sso_backup_dinamico(ws_com):
    """
    Una tabla = desde la fila con 'Id del incidente' hasta la fila anterior a la siguiente tabla
    (o fin de datos). Un solo rango contiguo por tabla: siempre incluye encabezados y recorta
    filas vacías/cero al final. Omite tablas donde todas las fechas son 0 (00-01-1900).
    """
    filas_h = _filas_encabezado_tablas_sso(ws_com)
    if not filas_h:
        state.errores.append(
            "[ERROR] SSO: no se encontró ninguna tabla con encabezado tipo 'Id del incidente'."
        )
        return []

    used = ws_com.UsedRange
    sheet_max_r = used.Row + used.Rows.Count - 1
    rangos = []

    for idx, (h_row, marcador_col) in enumerate(filas_h):
        next_h = filas_h[idx + 1][0] if idx + 1 < len(filas_h) else None
        max_row_bloque = (next_h - 1) if next_h else sheet_max_r
        if max_row_bloque < h_row:
            max_row_bloque = h_row

        # Columnas fijas: A (1) hasta J (10)
        min_c = 1
        last_c = 10
        last_row = _ultima_fila_con_datos_en_rango_com(ws_com, min_c, h_row, last_c, max_row_bloque)

        if not _tabla_sso_tiene_datos(ws_com, h_row, last_row, min_c, last_c):
            continue

        rango_a1 = (
            f"{get_column_letter(min_c)}{h_row}:"
            f"{get_column_letter(last_c)}{last_row}"
        )
        rangos.append(rango_a1)

    return rangos

# Exporta un rango SSO ocultando filas con ID = 0, operando directamente sobre ws_com.
def exportar_imagen_sso_filtrada(ruta_excel, ws_com, rango, nombre_imagen):
    import re

    carpeta_temp = r"C:\\Temp"
    if not os.path.exists(carpeta_temp):
        os.makedirs(carpeta_temp)
    imagen_salida = os.path.join(carpeta_temp, nombre_imagen)

    match = re.match(r'([A-Z]+)(\d+):([A-Z]+)(\d+)', rango)
    if not match:
        state.errores.append(f"[ERROR] SSO: rango inválido '{rango}'")
        return imagen_salida

    col_ini = column_index_from_string(match.group(1))
    row_ini = int(match.group(2))
    col_fin = column_index_from_string(match.group(3))
    row_fin = int(match.group(4))

    # Buscar columna "Id del incidente"
    col_id = None
    for c in range(col_ini, col_fin + 1):
        v = ws_com.Cells(row_ini, c).Value
        if v is not None and SSO_MARCADOR_TABLA in str(v).strip().lower():
            col_id = c
            break

    # Ocultar filas con ID = 0 o None
    filas_ocultas = []
    if col_id:
        for r in range(row_ini + 1, row_fin + 1):
            v = ws_com.Cells(r, col_id).Value
            if v is None or v == 0:
                ws_com.Rows(r).Hidden = True
                filas_ocultas.append(r)

    # Exportar via portapapeles (mapa de bits) para máxima calidad de imagen.
    try:
        rng = ws_com.Range(rango)
        rng.CopyPicture(Appearance=1, Format=2)
        time.sleep(1)
        img = ImageGrab.grabclipboard()
        if img:
            img.save(imagen_salida, "PNG")
        else:
            msg = f"[ERROR] SSO: no se pudo obtener imagen del portapapeles ({rango})"
            state.errores.append(msg)
            print(f"  ✗ {msg}")
    except Exception as e:
        msg = f"[ERROR] SSO: falló exportación {rango}: {e}"
        state.errores.append(msg)
        print(f"  ✗ {msg}")

    # Restaurar filas ocultas
    for r in filas_ocultas:
        ws_com.Rows(r).Hidden = False

    return imagen_salida

# Patrón de nombre de archivo esperado por faena (búsqueda case-insensitive).
_PATRON_EXCEL_FAENA = {
    "MLP":             "mlp semana",
    "CEN":             "informe semanal",
    "ANT":             "informe semanal",
    "CMZ":             "proyectado",
    "FCAB":            "amsa",
    "SSO":             "eventos seguridad",
    "Gestión Hídrica": "seguimiento",
}


# Fragmentos de nombre de archivo que NO deben actualizarse.
_VINCULOS_IGNORAR = [
    "cd mina",
]

def _buscar_excel_en_carpeta(carpeta, patron, clave):
    """Busca un .xlsx cuyo nombre contenga 'patron'. Si no lo encuentra, abre selector."""
    if os.path.isdir(carpeta):
        candidatos = [
            f for f in os.listdir(carpeta)
            if f.lower().endswith(".xlsx")
            and not f.startswith("~$")
            and patron in f.lower()
        ]
        if len(candidatos) == 1:
            return os.path.join(carpeta, candidatos[0])
        if len(candidatos) > 1:
            print(f"  ! {clave}: múltiples coincidencias para '{patron}' en {carpeta} → abriendo selector")
        else:
            print(f"  ! {clave}: no se encontró '{patron}' en {carpeta} → abriendo selector")
    else:
        print(f"  ! {clave}: carpeta no existe: {carpeta} → abriendo selector")

    return seleccionar_archivo(f"Excel de {clave}")

# Redirige los vínculos externos del workbook a las subcarpetas de faena de la semana actual.
def actualizar_vinculos_faenas(wb, informes_dirs, carpeta_raiz=None):  # carpeta_raiz reservado para uso futuro
    """
    Para cada vínculo externo:
    - Si el nombre contiene un fragmento de _VINCULOS_IGNORAR → se omite.
    - Si la subcarpeta de faena aparece en el path → busca por patrón en esa subcarpeta.
    - Si el nombre del archivo coincide con un patrón de _PATRON_EXCEL_RAIZ → busca en carpeta_raiz.
    - Si nada coincide → abre el explorador.
    """
    try:
        links = wb.LinkSources(1)  # 1 = xlExcelLinks
        if not links:
            print("  (sin vínculos externos)")
            return

        for link_origen in links:
            link_norm = link_origen.replace("/", "\\").lower()
            nombre_archivo = os.path.basename(link_origen).lower()

            # 1. Ignorar vínculos explícitamente excluidos
            if any(ignorar in nombre_archivo for ignorar in _VINCULOS_IGNORAR):
                print(f"  — omitido: {os.path.basename(link_origen)}")
                continue

            nueva_ruta = None
            clave_encontrada = None

            # 2. Buscar por subcarpeta de faena
            for clave, carpeta_nueva in informes_dirs.items():
                subcarpeta = os.path.basename(str(carpeta_nueva)).lower()
                if subcarpeta in link_norm:
                    patron = _PATRON_EXCEL_FAENA.get(clave, "")
                    nueva_ruta = _buscar_excel_en_carpeta(str(carpeta_nueva), patron, clave)
                    clave_encontrada = clave
                    break

            # 3. No reconocido → selector
            if clave_encontrada is None:
                print(f"  ? Vínculo no reconocido: {os.path.basename(link_origen)} → abriendo selector")
                nueva_ruta = seleccionar_archivo(f"Excel vinculado ({os.path.basename(link_origen)})")
                clave_encontrada = "?"

            if nueva_ruta and nueva_ruta.lower() != link_norm:
                try:
                    wb.ChangeLink(link_origen, nueva_ruta, 1)
                    print(f"  ✓ {clave_encontrada}: {os.path.basename(nueva_ruta)}")
                except Exception as e:
                    print(f"  ! No se pudo actualizar vínculo ({clave_encontrada}): {e}")


    except Exception as e:
        print(f"  ! Error general al procesar vínculos: {e}")

# Escribe las fechas de inicio y término de la semana en la hoja AUX del Excel madre.
def _ordenar_hoja_sso(wb):
    """Ordena cada tabla SSO por su columna Fecha de forma ascendente y permanente."""
    try:
        ws = wb.Worksheets("SSO")
        used = ws.UsedRange
        sheet_max_r = used.Row + used.Rows.Count - 1
        filas_h = [r for r, _ in _filas_encabezado_tablas_sso(ws)]
        if not filas_h:
            return
        tablas_ordenadas = 0
        for idx, h_row in enumerate(filas_h):
            next_h = filas_h[idx + 1] if idx + 1 < len(filas_h) else None
            last_row = (next_h - 1) if next_h is not None else sheet_max_r
            if last_row <= h_row:
                continue
            min_c = _columna_izquierda_tabla_sso(ws, h_row)
            last_c = _ultima_columna_cabecera_sso(ws, h_row)
            col_id = None
            col_fecha = None
            for c in range(min_c, last_c + 1):
                v = ws.Cells(h_row, c).Value
                if not v:
                    continue
                v_low = str(v).strip().lower()
                if col_id is None and SSO_MARCADOR_TABLA in v_low:
                    col_id = c
                if col_fecha is None and "fecha" in v_low:
                    col_fecha = c
            if not col_id and not col_fecha:
                continue
            data_range = ws.Range(ws.Cells(h_row, min_c), ws.Cells(last_row, last_c))
            if col_id and col_fecha:
                data_range.Sort(
                    Key1=ws.Cells(h_row, col_id),    Order1=1,
                    Key2=ws.Cells(h_row, col_fecha),  Order2=1,
                    Header=1, Orientation=1,
                )
            elif col_id:
                data_range.Sort(Key1=ws.Cells(h_row, col_id), Order1=1, Header=1, Orientation=1)
            else:
                data_range.Sort(Key1=ws.Cells(h_row, col_fecha), Order1=1, Header=1, Orientation=1)
            tablas_ordenadas += 1
        print(f"  ✓ SSO ordenada ({tablas_ordenadas} tabla(s) por fecha)")
    except Exception as e:
        msg = f"[REVISAR] No se pudo ordenar SSO: {e}"
        state.errores.append(msg)
        print(f"  ! {msg}")

def _buscar_wb_en_excel_usuario(ruta_excel):
    """
    Intenta conectarse al Excel ya abierto por el usuario (Dispatch, no DispatchEx)
    y devuelve el workbook si lo encuentra. No almacena nada en state.
    Retorna (wb, excel_app) o (None, None).
    """
    try:
        excel_usuario = win32.Dispatch("Excel.Application")
        basename = os.path.basename(ruta_excel).lower()
        count = excel_usuario.Workbooks.Count
        for i in range(1, count + 1):
            try:
                wb = excel_usuario.Workbooks(i)
                if os.path.basename(wb.FullName).lower() == basename:
                    return wb, excel_usuario
            except Exception:
                continue
    except Exception:
        pass
    return None, None


def _refrescar_todos_los_vinculos(wb):
    """
    Fuerza el refresco de TODOS los vínculos externos del workbook, de a uno.
    Actualizar todos juntos hace que un solo fallo aborte el lote completo y deje
    el COM inestable. Actualizando individualmente se aíslan los errores.
    """
    try:
        links = wb.LinkSources(1)   # 1 = xlExcelLinks
        if not links:
            return
        print(f"  Refrescando {len(links)} vínculo(s) externos…")
        errores_link = []
        for link in links:
            try:
                wb.UpdateLink(link, 1)
            except Exception as e:
                errores_link.append(os.path.basename(str(link)))
        if errores_link:
            print(f"  ! No se pudieron refrescar {len(errores_link)} vínculo(s): {', '.join(errores_link)}")
    except Exception as e:
        print(f"  ! Error al obtener lista de vínculos: {e}")


def _cerrar_wb_por_nombre(excel_app, ruta_objetivo, limpiar_cache=False):
    """Cierra el workbook cuyo nombre de archivo coincide con ruta_objetivo, si está abierto."""
    nombre_objetivo = os.path.basename(ruta_objetivo).lower()
    try:
        for i in range(excel_app.Workbooks.Count, 0, -1):
            try:
                wb_test = excel_app.Workbooks(i)
                if os.path.basename(wb_test.FullName).lower() == nombre_objetivo:
                    wb_test.Close(False)
                    if limpiar_cache:
                        # Eliminar del cache por ruta completa o nombre
                        for k in list(state._workbooks_abiertos.keys()):
                            if os.path.basename(str(k)).lower() == nombre_objetivo:
                                del state._workbooks_abiertos[k]
                    break
            except Exception:
                pass
    except Exception:
        pass


def abrir_excel_y_actualizar_vinculos(ruta_excel, informes_dirs, carpeta_destino=None, carpeta_raiz=None, ordenar_sso=False):
    """
    Actualiza vínculos y guarda una copia _act en carpeta_destino.
    Usa el Excel del usuario si está abierto (para tener datos ya calculados),
    o DispatchEx si no está abierto.
    """
    try:
        nombre_base = os.path.splitext(os.path.basename(ruta_excel))[0]
        ruta_act = os.path.join(carpeta_destino, f"{nombre_base}_act.xlsx") if carpeta_destino else None

        # Intentar primero con el Excel del usuario (evita problemas de read-only).
        # Hacerlo ANTES de _obtener_excel_app() para que Dispatch encuentre
        # el proceso del usuario y no nuestro DispatchEx.
        wb_usuario, excel_usuario = _buscar_wb_en_excel_usuario(ruta_excel)

        if wb_usuario is not None:
            print("  Usando Excel del usuario para actualizar vínculos.")
            excel_usuario.DisplayAlerts = False
            print("\nActualizando vínculos externos...")
            actualizar_vinculos_faenas(wb_usuario, informes_dirs, carpeta_raiz=carpeta_raiz)
            _refrescar_todos_los_vinculos(wb_usuario)
            if ordenar_sso:
                print("\nOrdenando tablas SSO por fecha...")
                _ordenar_hoja_sso(wb_usuario)
            print("  Recalculando fórmulas con vínculos actualizados...")
            try:
                excel_usuario.CalculateFull()
                time.sleep(2)
            except Exception as e:
                print(f"  ! No se pudo recalcular: {e}")
            if ruta_act:
                # Pausa breve para que el COM cross-proceso se estabilice tras el refresco
                time.sleep(1)
                # Cerrar _act si ya está abierto en la instancia del usuario
                _cerrar_wb_por_nombre(excel_usuario, ruta_act)
                try:
                    wb_usuario.SaveCopyAs(ruta_act)
                    print(f"  ✓ Copia guardada en: {ruta_act}")
                except Exception as e:
                    print(f"  ! No se pudo guardar copia _act: {e}")
            # NO guardar en _workbooks_abiertos: proxy cross-process rompe .Worksheets()
            return

        # El archivo no está abierto — usar DispatchEx.
        excel = _obtener_excel_app()
        wb = excel.Workbooks.Open(ruta_excel, UpdateLinks=0)
        state._workbooks_abiertos[ruta_excel] = wb

        print("\nActualizando vínculos externos...")
        actualizar_vinculos_faenas(wb, informes_dirs, carpeta_raiz=carpeta_raiz)
        _refrescar_todos_los_vinculos(wb)
        if ordenar_sso:
            print("\nOrdenando tablas SSO por fecha...")
            _ordenar_hoja_sso(wb)
        print("  Recalculando fórmulas con vínculos actualizados...")
        try:
            wb.Application.CalculateFull()
            time.sleep(2)
        except Exception as e:
            print(f"  ! No se pudo recalcular: {e}")
        if ruta_act:
            # Cerrar _act si ya está abierto en nuestra instancia DispatchEx
            _cerrar_wb_por_nombre(excel, ruta_act, limpiar_cache=True)
            try:
                wb.SaveCopyAs(ruta_act)
                print(f"  ✓ Copia guardada en: {ruta_act}")
            except Exception as e:
                print(f"  ! No se pudo guardar copia _act: {e}")

    except Exception as e:
        msg = f"[ERROR] No se pudieron actualizar los vínculos: {e}"
        state.errores.append(msg)
        print(f"  ✗ {msg}")

def escribir_fechas_excel(ruta_excel, dia_inicio, mes_inicio, dia_fin, mes_fin):
    try:
        excel = _obtener_excel_app()
        wb = state._workbooks_abiertos.get(ruta_excel)
        if wb is None:
            wb = excel.Workbooks.Open(ruta_excel, UpdateLinks=0)
            state._workbooks_abiertos[ruta_excel] = wb

        ws = wb.Worksheets("AUX")
        ws.Range("B4").Value = int(dia_inicio)
        ws.Range("C4").Value = int(mes_inicio)
        ws.Range("B5").Value = int(dia_fin)
        ws.Range("C5").Value = int(mes_fin)
        wb.Save()
    except Exception as e:
        state.errores.append(f"[ERROR] No se pudieron escribir fechas en AUX: {e}")

# Extrae el texto resumen desde el Excel base.
def extraer_resumen_excel(ruta_excel):
    try:
        wb = openpyxl.load_workbook(ruta_excel, data_only=True)
        sheet = wb["Grupo Minero FCAB PLAN"]
        resumen = []
        for row in range(38, 44):
            val = sheet[f"B{row}"].value
            if val:
                resumen.append(str(val))
        return "\n".join(resumen)
    except Exception as e:
        state.errores.append(f"[ERROR] No se pudo extraer resumen del Excel madre: {e}")
        return "Resumen no disponible."
